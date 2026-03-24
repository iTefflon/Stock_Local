#!/usr/bin/env python3
"""
Consolidación de Stock Físico - Tecnobox Sevilla
Suma el stock físico de productos con el mismo código y calcula diferencias con ERP
"""

import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─── Leer archivo ODS ───────────────────────────────────────────────
print("Leyendo Contro_Stock_Sevilla.ods ...")
df_raw = pd.read_excel(
    'Contro_Stock_Sevilla.ods',
    sheet_name='Hoja1',
    engine='odf',
    header=None
)

df_raw.columns = ['CODIGO', 'PRODUCTO', 'MARCA', 'CATEGORIA',
                  'STOCK_FISICO', 'STOCK_ERP', 'PRECIO_UNITARIO']

# Eliminar fila de encabezado y filas vacías
df = df_raw[df_raw['CODIGO'] != 'CODIGO'].copy()
df = df.dropna(subset=['CODIGO']).copy()
df = df[df['CODIGO'].astype(str).str.strip() != ''].copy()

# Limpiar y convertir tipos
def to_num(val):
    try:
        return float(str(val).replace('.', '').replace(',', '.'))
    except:
        return 0

df['STOCK_FISICO'] = df['STOCK_FISICO'].apply(to_num)
df['STOCK_ERP']    = df['STOCK_ERP'].apply(to_num)
df['PRECIO_UNITARIO'] = df['PRECIO_UNITARIO'].apply(to_num)
df['CODIGO'] = df['CODIGO'].astype(str).str.strip()

print(f"Filas cargadas: {len(df)}")
print(f"Códigos únicos: {df['CODIGO'].nunique()}")

# ─── Consolidar: sumar STOCK_FISICO por CODIGO ───────────────────────
# Para la consolidación tomamos el primer valor de nombre/marca/cat/precio
df_consol = df.groupby('CODIGO', sort=False).agg(
    PRODUCTO      = ('PRODUCTO',       'first'),
    MARCA         = ('MARCA',          'first'),
    CATEGORIA     = ('CATEGORIA',      'first'),
    STOCK_FISICO  = ('STOCK_FISICO',   'sum'),
    STOCK_ERP     = ('STOCK_ERP',      'first'),
    PRECIO_UNITARIO = ('PRECIO_UNITARIO', 'first'),
    VECES_CONTADO = ('STOCK_FISICO',   'count'),
).reset_index()

df_consol['DIFERENCIA'] = df_consol['STOCK_FISICO'] - df_consol['STOCK_ERP']
df_consol['STOCK_FISICO']  = df_consol['STOCK_FISICO'].astype(int)
df_consol['STOCK_ERP']     = df_consol['STOCK_ERP'].astype(int)
df_consol['DIFERENCIA']    = df_consol['DIFERENCIA'].astype(int)

print(f"\nProductos consolidados: {len(df_consol)}")
print(f"Códigos que se repitieron: {(df_consol['VECES_CONTADO'] > 1).sum()}")
print(f"Con diferencia positiva (sobrante): {(df_consol['DIFERENCIA'] > 0).sum()}")
print(f"Con diferencia negativa (faltante): {(df_consol['DIFERENCIA'] < 0).sum()}")
print(f"Sin diferencia:                     {(df_consol['DIFERENCIA'] == 0).sum()}")

# ─── Exportar a Excel con formato ────────────────────────────────────
output_file = 'Stock_Consolidado_Sevilla.xlsx'

cols_order = ['CODIGO', 'PRODUCTO', 'MARCA', 'CATEGORIA',
              'STOCK_FISICO', 'STOCK_ERP', 'DIFERENCIA',
              'PRECIO_UNITARIO', 'VECES_CONTADO']

df_export = df_consol[cols_order].copy()
df_export.to_excel(output_file, index=False, sheet_name='Stock_Consolidado')

# ─── Aplicar formato con openpyxl ────────────────────────────────────
wb = load_workbook(output_file)
ws = wb.active

# Colores
COLOR_HEADER_BG  = '276221'   # verde oscuro tecnobox
COLOR_HEADER_FG  = 'FFFFFF'
COLOR_ROJO_BG    = 'FDDCDC'   # faltante
COLOR_VERDE_BG   = 'D4EDDA'   # sobrante
COLOR_GRIS_BG    = 'F5F5F5'   # sin diferencia
COLOR_ALT_BG     = 'F9F9F9'

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Ancho de columnas
col_widths = {
    'A': 20,   # CODIGO
    'B': 55,   # PRODUCTO
    'C': 18,   # MARCA
    'D': 25,   # CATEGORIA
    'E': 14,   # STOCK_FISICO
    'F': 12,   # STOCK_ERP
    'G': 14,   # DIFERENCIA
    'H': 17,   # PRECIO_UNITARIO
    'I': 16,   # VECES_CONTADO
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Encabezados
header_labels = [
    'Código', 'Producto', 'Marca', 'Categoría',
    'Stock Físico', 'Stock ERP', 'Diferencia',
    'Precio Unit.', 'Veces Contado'
]
for col_idx, label in enumerate(header_labels, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = label
    cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
    cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

ws.row_dimensions[1].height = 28

# Filas de datos
for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    diferencia = ws.cell(row=row_idx, column=7).value or 0
    if diferencia < 0:
        row_bg = COLOR_ROJO_BG
    elif diferencia > 0:
        row_bg = COLOR_VERDE_BG
    else:
        row_bg = COLOR_GRIS_BG if row_idx % 2 == 0 else COLOR_ALT_BG

    for col_idx, cell in enumerate(row, 1):
        cell.fill = PatternFill(fill_type='solid', fgColor=row_bg)
        cell.border = border
        cell.alignment = Alignment(
            horizontal='center' if col_idx in (1, 5, 6, 7, 8, 9) else 'left',
            vertical='center',
            wrap_text=(col_idx == 2)
        )
        if col_idx == 7:  # Diferencia: negrita y color
            if diferencia < 0:
                cell.font = Font(bold=True, color='C0392B')
            elif diferencia > 0:
                cell.font = Font(bold=True, color='1A7A31')
            else:
                cell.font = Font(bold=False, color='555555')

# Auto-filter
ws.auto_filter.ref = ws.dimensions

# Congelar primera fila
ws.freeze_panes = 'A2'

wb.save(output_file)
print(f"\n✅ Archivo generado: {output_file}")

# ─── Exportar JSON para la página web ────────────────────────────────
json_data = []
for _, row in df_consol.iterrows():
    diff = int(row['DIFERENCIA'])
    if diff < 0:
        estado = 'faltante'
    elif diff > 0:
        estado = 'sobrante'
    else:
        estado = 'ok'
    json_data.append({
        'codigo':    str(row['CODIGO']),
        'producto':  str(row['PRODUCTO']),
        'marca':     str(row['MARCA']),
        'categoria': str(row['CATEGORIA']),
        'stockFisico': int(row['STOCK_FISICO']),
        'stockERP':    int(row['STOCK_ERP']),
        'diferencia':  diff,
        'precio':      int(row['PRECIO_UNITARIO']) if row['PRECIO_UNITARIO'] else 0,
        'vecesContado': int(row['VECES_CONTADO']),
        'estado':      estado
    })

os.makedirs('web', exist_ok=True)
with open('web/stock_data.json', 'w', encoding='utf-8') as f:
    json.dump(json_data, f, ensure_ascii=False, indent=2)

print(f"✅ JSON generado: web/stock_data.json ({len(json_data)} productos)")
